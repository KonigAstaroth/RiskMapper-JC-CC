from django.http import HttpResponse
from django.conf import settings
from django.shortcuts import redirect
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from app.src.library_service import searchEvent
from datetime import datetime, timezone
import datetime
import os

def wrap(ws, celda, valor):
    ws[celda] = valor
    ws[celda].alignment = Alignment(wrap_text=True, vertical="top")

def getPlace(filters):
    location = ""
    estado = filters.get('Estados_hechos')
    municipio = filters.get('Municipio_hechos')
    if estado:
        location += f"{estado} -"
    if municipio:
        location += f"{municipio} -"

    return location


def downloadEvents(request):
    filters = request.session.get('filters_library')
    events = searchEvent(filters)

    if not events or not filters:
        return redirect('library')
    
    location = getPlace(filters)

    file_path = os.path.join(settings.BASE_DIR, 'app', 'xlsx_templates', 'events.xlsx')
    wb = load_workbook(file_path)
    events_row = 2 
    ws_events = wb['Eventos']

    for event in events:

        date = event.get('FechaHoraHecho')

        if date:
            date_str = date.strftime("%d-%B-%Y")
            hour_str = date.strftime("%I:%M %p")
        else:
            date_str = ""
            hour_str = ""
        wrap(ws_events, f"A{events_row}", event.get('Delito'))
        wrap(ws_events, f"B{events_row}", event.get('Calle_hechos'))
        wrap(ws_events, f"C{events_row}", event.get('ColoniaHechos'))
        wrap(ws_events, f"D{events_row}", event.get('Municipio_hechos'))
        wrap(ws_events, f"E{events_row}", event.get('Estado_hechos'))
        wrap(ws_events, f"F{events_row}", date_str)
        wrap(ws_events, f"G{events_row}", hour_str)

        events_row += 1
    
    now = datetime.datetime.now(timezone.utc)
    month = now.month
    year = now.year

    months = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Eventos - {location} {months[month-1]} - {year}.xlsx"'
    )

    
    wb.save(response)
    return response